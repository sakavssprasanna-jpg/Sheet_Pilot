import io
import pandas as pd
from typing import Optional

def to_csv(df: pd.DataFrame) -> bytes:
    """
    Convert a Pandas DataFrame into CSV bytes for user download.
    
    Args:
        df: The pandas DataFrame.
        
    Returns:
        CSV encoded bytes.
    """
    if df is None:
        return b""
    return df.to_csv(index=False).encode("utf-8")

def to_excel(df: pd.DataFrame) -> bytes:
    """
    Convert a Pandas DataFrame into a styled Excel (XLSX) bytes for user download.
    
    Args:
        df: The pandas DataFrame.
        
    Returns:
        XLSX formatted bytes.
    """
    if df is None:
        return b""
    
    output = io.BytesIO()
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="SheetPilot_Export")
            workbook = writer.book
            worksheet = writer.sheets["SheetPilot_Export"]
            
            # Apply styling to headers (slate/dark blue theme matching SheetPilot)
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0E5A84", end_color="0E5A84", fill_type="solid") # Deep teal/blue accent
            
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Auto-adjust column widths
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 10)
                
        return output.getvalue()
    except Exception:
        return b""
